"""
generate_report_latex.py
Reads an inspection Excel export and produces a branded PDF via XeLaTeX.
Charts are rendered as PNGs by matplotlib and included via \\includegraphics.

Usage:
    uv run --with pandas --with openpyxl generate_report_latex.py <input.xlsx> [output.pdf]
"""

import sys
import re
import json
import base64
import shutil
import tempfile
import zipfile
import subprocess
import xml.etree.ElementTree as ET
from pathlib import Path
from collections import defaultdict

import pandas as pd


# ---------------------------------------------------------------------------
# XeLaTeX discovery
# ---------------------------------------------------------------------------

XELATEX_CANDIDATES = [
    # Portable MiKTeX bundled alongside this script (gitignored)
    str(Path(__file__).parent / "miktex" / "texmfs" / "install" / "miktex" / "bin" / "x64" / "xelatex.exe"),
    # System-wide installs
    "xelatex",
    r"C:\Users\user\AppData\Local\Programs\MiKTeX\miktex\bin\x64\xelatex.exe",
    r"C:\Program Files\MiKTeX\miktex\bin\x64\xelatex.exe",
    r"C:\Program Files (x86)\MiKTeX\miktex\bin\xelatex.exe",
]

def find_xelatex() -> str:
    for candidate in XELATEX_CANDIDATES:
        if shutil.which(candidate) or Path(candidate).exists():
            return candidate
    raise RuntimeError(
        "xelatex not found. Install MiKTeX and ensure it is on PATH."
    )


# ---------------------------------------------------------------------------
# Brand constants
# ---------------------------------------------------------------------------

NAVY   = "#25408F"
SKY    = "#00B0F0"
GRAY   = "#F0F3F6"
GREEN  = "#85CF5F"
AMBER  = "#F0B557"
RED    = "#F05773"
SLATE  = "#4D5B82"


def score_color(score: float) -> str:
    if score >= 85: return GREEN
    if score >= 80: return AMBER
    return RED


def hex_to_rgb(h: str) -> tuple:
    h = h.lstrip("#")
    return tuple(int(h[i:i+2], 16) / 255 for i in (0, 2, 4))


# ---------------------------------------------------------------------------
# Company config  (gitignored — same as HTML version)
# ---------------------------------------------------------------------------

def load_company(script_path: Path) -> dict:
    root = script_path.parent
    company = {"name": "", "address": "", "phone": "", "logo_path": ""}
    config = root / "company.json"
    if config.exists():
        with open(config, encoding="utf-8") as f:
            company.update(json.load(f))
    for ext in ("png", "jpg", "jpeg", "svg"):
        p = root / f"logo.{ext}"
        if p.exists():
            company["logo_path"] = str(p)
            break
    return company

COMPANY = load_company(Path(__file__))


# ---------------------------------------------------------------------------
# Data loading  (identical logic to HTML version)
# ---------------------------------------------------------------------------

def load_inspections(file_path: str):
    filters_df = pd.read_excel(file_path, sheet_name="Filters", header=None, engine="openpyxl")
    filters = dict(zip(filters_df[0].str.strip(), filters_df[1]))
    df = pd.read_excel(file_path, sheet_name="Raw Data", engine="openpyxl")
    complete = df[df["Status"].fillna("").str.strip().str.lower() == "complete"].copy()
    complete.reset_index(drop=True, inplace=True)
    return complete, filters, df


def comment_col(df: pd.DataFrame) -> str:
    return "Comments" if "Comments" in df.columns else "Comment"


def parse_location(location: str) -> tuple:
    location = str(location).strip()
    for sep in (" - ", "_ "):
        if sep in location:
            _, space = location.rsplit(sep, 1)
            return "", "", space.strip()
    return "", "", location


def drop_image_stubs(df: pd.DataFrame) -> pd.DataFrame:
    stub_pattern = re.compile(r"^\s*\(([^)]+)\)\s*$")
    def is_stub(comment, group_comments):
        m = stub_pattern.match(comment)
        if not m:
            return False
        tag = m.group(1)
        return any(
            c != comment and re.search(r"\(" + re.escape(tag) + r"\)\s*$", c)
            for c in group_comments
        )
    drop_idx = set()
    cc = comment_col(df)
    for _, group in df.groupby(["Inspection #", "Element"], sort=False):
        comments = group[cc].fillna("").tolist()
        for idx, comment in zip(group.index, comments):
            if is_stub(comment, comments):
                drop_idx.add(idx)
    return df[~df.index.isin(drop_idx)].reset_index(drop=True)


def build_summary(df: pd.DataFrame, filters: dict) -> dict:
    per_insp = (
        df.groupby("Inspection #", sort=False)
        .agg(
            score=("Score In %", "first"),
            location=("Location", "first"),
            completion=("Completion Date", "first"),
            completed_by=("Completed By", "first"),
            venue_=("Venue", "first"),
            building_=("Building", "first"),
            corporation_=("Corporation", "first"),
        )
        .reset_index()
    )

    overall_score = round(per_insp["score"].mean(), 2) if not per_insp.empty else 0
    venue = per_insp["venue_"].iloc[0] if not per_insp.empty else "Unknown Venue"
    building = per_insp["building_"].iloc[0] if not per_insp.empty else ""
    corporation = per_insp["corporation_"].iloc[0] if not per_insp.empty else ""

    date_from = pd.to_datetime(filters.get("From Date"), errors="coerce")
    date_to   = pd.to_datetime(filters.get("To Date"),   errors="coerce")
    if pd.notna(date_from) and pd.notna(date_to):
        if date_from.date() == date_to.date():
            date_str = date_from.strftime("%B %d, %Y")
        else:
            date_str = f"{date_from.strftime('%B %d, %Y')} – {date_to.strftime('%B %d, %Y')}"
    else:
        date_str = ""

    per_insp["_space"] = per_insp["location"].apply(lambda l: parse_location(str(l))[2])
    space_scores = per_insp.groupby("_space")["score"].mean().round(2).sort_values(ascending=True)

    insp_zone = (
        df.groupby("Inspection #")["Zone"].first()
        .reset_index().rename(columns={"Zone": "zone"})
    )
    per_insp_z = per_insp.merge(insp_zone, on="Inspection #", how="left")
    zone_scores = per_insp_z.groupby("zone")["score"].mean().round(2).sort_values(ascending=True)
    element_scores = (
        df.groupby(["Element", "Inspection #"])["Rating"].mean()
          .groupby("Element").mean()
          .round(2)
          .dropna()
          .sort_values(ascending=True)
    )

    _cc = comment_col(df)
    work_order_rows = df[df[_cc].notna() & (df[_cc].astype(str).str.strip() != "")].copy()
    work_order_rows = work_order_rows.rename(columns={_cc: "Comment"})
    work_order_rows = drop_image_stubs(work_order_rows)

    insp_meta = {
        row["Inspection #"]: {
            "completion": pd.to_datetime(row["completion"], errors="coerce"),
            "completed_by": str(row["completed_by"]).strip() if pd.notna(row["completed_by"]) else "",
        }
        for _, row in per_insp.iterrows()
    }

    return {
        "venue": venue, "building": building, "corporation": corporation,
        "date_str": date_str, "overall_score": overall_score,
        "total_inspections": len(per_insp),
        "total_location_types": len(space_scores),
        "total_elements": int(df["Element"].nunique()),
        "loc_scores": space_scores,
        "zone_scores": zone_scores,
        "element_scores": element_scores,
        "work_order_rows": work_order_rows,
        "insp_scores": dict(zip(per_insp["Inspection #"], per_insp["score"])),
        "insp_meta": insp_meta,
    }


def extract_images(file_path: str, raw_df: pd.DataFrame) -> dict:
    ns = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a":   "http://schemas.openxmlformats.org/drawingml/2006/main",
    }
    result = defaultdict(list)
    with zipfile.ZipFile(file_path) as z:
        names = z.namelist()
        drawing = next((n for n in names if n == "xl/drawings/drawing1.xml"), None)
        rels_path = next((n for n in names if n == "xl/drawings/_rels/drawing1.xml.rels"), None)
        if not drawing or not rels_path:
            return dict(result)
        rels_root = ET.fromstring(z.read(rels_path))
        rid_to_target = {r.attrib["Id"]: r.attrib["Target"] for r in rels_root}
        drawing_root = ET.fromstring(z.read(drawing))
        for anchor in drawing_root.findall("xdr:twoCellAnchor", ns):
            row_el = anchor.find("xdr:from/xdr:row", ns)
            blip = anchor.find(".//xdr:blipFill/a:blip", ns)
            if row_el is None or blip is None:
                continue
            raw_idx = int(row_el.text) - 1  # xdr:row=0 is header; data starts at 1
            rid = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
            if rid is None or raw_idx < 0 or raw_idx >= len(raw_df):
                continue
            insp_id = str(raw_df.iloc[raw_idx].get("Inspection #", ""))
            if not insp_id:
                continue
            img_rel = rid_to_target.get(rid, "")
            img_path = "xl/media/" + Path(img_rel).name
            if img_path not in names:
                continue
            result[insp_id].append(z.read(img_path))
    return dict(result)


def extract_image_links(file_path: str, max_per_insp: int = 6) -> dict:
    """Read attachment image URLs (cell hyperlinks) keyed by Inspection #.

    Some exports embed images; others put a filename in the Attachment column
    with the real image behind a cell hyperlink. This reads those links.
    """
    import openpyxl
    result = defaultdict(list)
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if "Raw Data" not in wb.sheetnames:
        return dict(result)
    ws = wb["Raw Data"]
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    att_col  = headers.get("Attachment")
    insp_col = headers.get("Inspection #")
    if not att_col or not insp_col:
        return dict(result)
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=att_col)
        url = cell.hyperlink.target if cell.hyperlink else None
        if not url and isinstance(cell.value, str) and cell.value.startswith("http"):
            url = cell.value
        if not url:
            continue
        insp_id = str(ws.cell(row=r, column=insp_col).value or "").strip()
        if insp_id and len(result[insp_id]) < max_per_insp:
            result[insp_id].append(url)
    return dict(result)


def download_images(links_by_insp: dict, log=print, timeout: int = 15, workers: int = 8) -> dict:
    """Download linked images concurrently. Returns bytes keyed by Inspection #."""
    import urllib.request
    from concurrent.futures import ThreadPoolExecutor

    tasks = [(insp_id, url) for insp_id, urls in links_by_insp.items() for url in urls]
    if not tasks:
        return {}

    def fetch(task):
        insp_id, url = task
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return insp_id, resp.read(), None
        except Exception as exc:  # network error, 404, timeout, etc.
            return insp_id, None, str(exc)

    result = defaultdict(list)
    ok = 0
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for insp_id, data, err in ex.map(fetch, tasks):
            if data:
                result[insp_id].append(data)
                ok += 1
            else:
                log(f"    [warn] image download failed: {err}")
    log(f"    Downloaded {ok}/{len(tasks)} linked images.")
    return dict(result)


# ---------------------------------------------------------------------------
# Chart rendering  (native TikZ — no external images)
# ---------------------------------------------------------------------------

def _bar_color_name(score: float) -> str:
    if score >= 85: return "bargreen"
    if score >= 80: return "barambr"
    return "barred"

def render_bar_chart_tex(scores: pd.Series, title: str, n: int = 5) -> str:
    n_total = len(scores)
    if n_total > n:
        # Top N: best scores descending (best at top)
        # Bottom N: worst scores reversed (worst at very bottom)
        top_n    = scores.iloc[-n:][::-1]
        bot_n    = scores.iloc[:n][::-1]
        display  = pd.concat([top_n, bot_n])
        n_hidden = max(0, n_total - n * 2)
    else:
        display  = scores[::-1]   # all bars, best at top
        n_hidden = 0

    labels = [str(l) for l in display.index]
    values = display.values.tolist()

    group_w = 0.55  # cm for the Top/Bottom group label column (only when truncated)
    label_w = 4.5   # cm for bar labels
    score_w = 1.4   # cm for score text after bar
    row_sep = 0.52  # cm between bar centres
    bar_h   = 0.30  # cm bar height
    sep_gap = row_sep * 1.6  # extra vertical room for the ellipsis row

    # Total label reservation: add group_w only when showing top/bottom groups
    total_label_w = label_w + (group_w if n_total > n else 0)

    rows = []
    for i, (label, value) in enumerate(zip(labels, values)):
        in_bottom = n_total > n and i >= n
        y = -i * row_sep - (sep_gap if in_bottom else 0)

        # Ellipsis separator between top-N and bottom-N groups (only when entries are hidden)
        if n_hidden > 0 and i == n:
            sep_y = y + sep_gap * 0.5
            rows.append(
                rf"  \node[font=\small\itshape, text=gray!70] at ({{0.5\barw}},{sep_y:.3f}cm)"
                rf" {{\ldots\ {n_hidden} more\ \ldots}};"
            )

        color = _bar_color_name(value)
        half  = bar_h / 2
        frac  = value / 100.0
        esc   = tex(label)
        rows.append(
            rf"  \node[anchor=east,font=\scriptsize,text width={label_w - 0.1:.1f}cm,align=right]"
            rf" at (0,{y:.3f}cm) {{{esc}}};"
        )
        rows.append(
            rf"  \fill[{color}] (0,{y - half:.3f}cm) rectangle ({{{frac:.4f}\barw}},{y + half:.3f}cm);"
        )
        rows.append(
            rf"  \node[anchor=west,font=\scriptsize\bfseries,text=navy]"
            rf" at ({{{frac:.4f}\barw}},{y:.3f}cm) {{\,{value:.2f}\%}};"
        )

    # Rotated group labels to the left of the bar labels
    if n_total > n:
        sep_offset   = sep_gap if n_hidden > 0 else 0
        group_x      = -(label_w + group_w * 0.5)
        top_center_y = -(n - 1) / 2.0 * row_sep
        bot_first_y  = -n * row_sep - sep_offset
        bot_center_y = bot_first_y - (n - 1) / 2.0 * row_sep
        rows.append(
            rf"  \node[rotate=90,anchor=center,font=\footnotesize\bfseries,text=slate]"
            rf" at ({group_x:.2f}cm,{top_center_y:.3f}cm) {{TOP {n}}};"
        )
        rows.append(
            rf"  \node[rotate=90,anchor=center,font=\footnotesize\bfseries,text=slate]"
            rf" at ({group_x:.2f}cm,{bot_center_y:.3f}cm) {{BOTTOM {n}}};"
        )

    n_display = len(labels)
    extra    = sep_gap if n_total > n else 0
    top_y    =  bar_h / 2
    bottom_y = -(n_display - 1) * row_sep - extra - bar_h / 2
    goal_line = [
        rf"  \draw[gray!80, dashed, line width=1.0pt]"
        rf" ({{0.80\barw}},{top_y:.3f}cm) -- ({{0.80\barw}},{bottom_y:.3f}cm);",
        rf"  \node[above, font=\scriptsize\bfseries, text=gray!80]"
        rf" at ({{0.80\barw}},{top_y:.3f}cm) {{Goal}};",
    ]
    return "\n".join([
        r"\noindent\begin{minipage}{\linewidth}",
        rf"\noindent\textbf{{\small\textcolor{{navy}}{{{title}}}}}",
        r"\par\vspace{3pt}",
        r"\noindent\begin{tikzpicture}",
        rf"  \setlength{{\barw}}{{\dimexpr\linewidth-{total_label_w:.2f}cm-{score_w:.1f}cm\relax}}",
        *rows,
        *goal_line,
        r"\end{tikzpicture}",
        r"\end{minipage}",
    ])


# ---------------------------------------------------------------------------
# LaTeX escaping
# ---------------------------------------------------------------------------

def tex(s: str) -> str:
    """Escape a string for LaTeX."""
    if not isinstance(s, str):
        s = str(s)
    replacements = [
        ("\\", "\\textbackslash{}"),
        ("&",  "\\&"),
        ("%",  "\\%"),
        ("$",  "\\$"),
        ("#",  "\\#"),
        # allow a line break after underscores so long tokens like "Foo_Common"
        # can wrap inside narrow fixed-width table columns
        ("_",  "\\_\\allowbreak "),
        ("{",  "\\{"),
        ("}",  "\\}"),
        ("~",  "\\textasciitilde{}"),
        ("^",  "\\textasciicircum{}"),
        ("–", "--"),
        ("—", "---"),
        ("’", "'"),
        ("“", "``"),
        ("”", "''"),
    ]
    for old, new in replacements:
        s = s.replace(old, new)
    return s


def color_cmd(hex_color: str) -> str:
    r, g, b = [int(hex_color.lstrip("#")[i:i+2], 16) for i in (0, 2, 4)]
    return f"{r},{g},{b}"


# ---------------------------------------------------------------------------
# LaTeX document builder
# ---------------------------------------------------------------------------

def fwdslash(p) -> str:
    """Convert a path to forward-slash form safe for LaTeX."""
    return str(p).replace("\\", "/")


def _md_inline(text: str) -> str:
    """Convert inline Markdown spans to LaTeX. tex() is applied to plain text segments only."""
    pattern = re.compile(r'`([^`]+)`|\*\*([^*]+)\*\*|\*([^*]+)\*')
    result = []
    last = 0
    for m in pattern.finditer(text):
        if m.start() > last:
            result.append(tex(text[last:m.start()]))
        code, bold, italic = m.group(1), m.group(2), m.group(3)
        if code is not None:
            result.append(rf'\colorbox{{gray!15}}{{\texttt{{{tex(code)}}}}}')
        elif bold is not None:
            result.append(rf'\textbf{{{tex(bold)}}}')
        elif italic is not None:
            result.append(rf'\textit{{{tex(italic)}}}')
        last = m.end()
    if last < len(text):
        result.append(tex(text[last:]))
    return "".join(result)


def md_to_tex(text: str) -> str:
    """Convert a Markdown subset to LaTeX.

    Supported: bullet lists (-, *, +), inline code, **bold**, *italic*, blank-line paragraphs.
    tex() is applied to plain text segments after Markdown tokens are extracted,
    so Markdown symbols are never pre-escaped into LaTeX commands.
    """
    lines = text.splitlines()
    out = []
    in_list = False

    for line in lines:
        stripped = line.strip()
        is_bullet = len(stripped) >= 2 and stripped[0] in "-*+" and stripped[1] == " "

        if is_bullet:
            if not in_list:
                out.append(r"\begin{itemize}")
                in_list = True
            out.append(rf"  \item {_md_inline(stripped[2:])}")
        else:
            if in_list:
                out.append(r"\end{itemize}")
                in_list = False
            if stripped == "":
                out.append(r"\par\vspace{4pt}")
            elif stripped.startswith("## "):
                out.append(rf"\par\noindent\textbf{{\small {_md_inline(stripped[3:])}}}\par\nopagebreak")
            elif stripped.startswith("# "):
                out.append(rf"\par\noindent\textbf{{\normalsize {_md_inline(stripped[2:])}}}\par\nopagebreak")
            else:
                out.append(_md_inline(stripped))

    if in_list:
        out.append(r"\end{itemize}")

    return "\n".join(out)


def build_comments_tex(comments: dict) -> str:
    """Render optional cover-page comments block. Only sections with content are shown."""
    LABELS = [
        ("client",  "Client Comments"),
        ("maintenance", "Maintenance Notes"),
        ("general", "General Remarks"),
    ]
    sections = [(label, comments[key]) for key, label in LABELS if comments.get(key, "").strip()]
    if not sections:
        return ""

    rows = []
    for i, (label, body) in enumerate(sections):
        if i > 0:
            rows.append(r"\par\vspace{6pt}")
        rows.append(rf"\noindent\textbf{{\small\textcolor{{navy}}{{\MakeUppercase{{{tex(label)}}}}}}}\\[2pt]")
        rows.append(rf"{{\small\color{{slate}}{{{md_to_tex(body)}}}}}")
        rows.append(r"\par")

    return "\n".join([
        r"\vspace{0.5cm}",
        r"\noindent\colorbox{statbg}{\parbox{\dimexpr\linewidth-2\fboxsep}{\vspace{8pt}",
        *rows,
        r"\vspace{8pt}}}",
    ])


def build_latex(summary: dict, zone_chart: str, loc_chart: str, elem_chart: str,
                insp_images: dict, tmp_dir: Path, comments: dict = None) -> str:

    venue       = tex(summary["venue"])
    date_str    = tex(summary["date_str"])
    overall     = summary["overall_score"]
    logo_path   = COMPANY.get("logo_path", "")
    company_name = tex(COMPANY.get("name", ""))

    # Fixed-width p-columns (no tabularx X). Two reasons:
    #   1. tabularx miscomputes X width when the table contains \multicolumn rows
    #      (our inspection-header/score rows) — known incompatibility.
    #   2. All-p-columns let \rowcolor fill every tabcolsep uniformly, so the
    #      header bar has no uncolored gap (the X column's left gap was the bug).
    # Widths sum to 15.1cm; with 10*tabcolsep (1.4cm) the table is ~16.5cm < the
    # 17.59cm textwidth. Kept in COL_W so the colspec and the full-span inspection
    # rows stay in sync.
    COL_W = {"zone": 2.7, "loc": 2.7, "elem": 2.7, "rating": 1.3, "comment": 6.7}
    colspec = (
        rf"L{{{COL_W['zone']}cm}} L{{{COL_W['loc']}cm}} L{{{COL_W['elem']}cm}} "
        rf"R{{{COL_W['rating']}cm}} L{{{COL_W['comment']}cm}}"
    )
    # Width a \multicolumn{5} cell spans: all column widths + the 4 internal
    # column gaps (each 2*tabcolsep).
    span_w = rf"\dimexpr {sum(COL_W.values()):.1f}cm+8\tabcolsep\relax"

    thead = (
        r"\rowcolor{theadrow}"
        r"\textcolor{white}{\textbf{Zone}} &"
        r"\textcolor{white}{\textbf{Location Type}} &"
        r"\textcolor{white}{\textbf{Element}} &"
        r"\textcolor{white}{\textbf{Rating}} &"
        r"\textcolor{white}{\textbf{Comments}}"
        r" \\\hline"
        "\n"
        r"\endhead"
    )

    overall_color = color_cmd(score_color(overall))

    # --- Stat cards (4 minipages) ---
    def stat_card(value, label):
        return (
            r"\begin{minipage}[t]{0.23\textwidth}"
            r"\centering"
            r"\colorbox{statbg}{\parbox{\dimexpr\linewidth-2\fboxsep}{"
            r"\hyphenpenalty=10000\exhyphenpenalty=10000"
            r"\parbox[c][2.4cm][c]{\linewidth}{\centering"
            rf"\textbf{{\LARGE\textcolor{{navy}}{{{value}}}}}\\[6pt]"
            rf"\textbf{{\small\textcolor{{slate}}{{\MakeUppercase{{{label}}}}}}}"
            r"}"
            r"}}"
            r"\end{minipage}"
        )

    cards = r"\hfill".join([
        stat_card(summary["total_inspections"],    "Inspections completed"),
        stat_card(summary["total_location_types"], "Location types inspected"),
        stat_card(summary["total_elements"],       "Unique elements inspected"),
        stat_card(len(summary["work_order_rows"]), "Deficiencies flagged"),
    ])

    highlights = build_comments_tex(comments or {})

    # --- Work orders table rows ---
    work_rows = summary["work_order_rows"]
    insp_scores = summary["insp_scores"]
    insp_meta   = summary["insp_meta"]

    table_rows = []
    current_insp = None

    for _, row in work_rows.iterrows():
        insp_id  = row.get("Inspection #", "")
        location = str(row.get("Location", ""))
        _, _, space = parse_location(location)
        zone    = str(row.get("Zone", "") or "").strip()
        element = str(row.get("Element", "") or "")
        comment = str(row.get("Comment", "") or "").strip()
        rating_raw = row.get("Rating", None)
        try:
            rating = f"{int(float(rating_raw))}\\%" if pd.notna(rating_raw) else ""
        except (ValueError, TypeError):
            rating = tex(str(rating_raw))

        if insp_id != current_insp:
            current_insp = insp_id
            insp_score = insp_scores.get(insp_id, "")
            score_txt  = f"{insp_score}\\%" if insp_score != "" else ""
            score_rgb  = color_cmd(score_color(float(insp_score))) if insp_score != "" else color_cmd(SLATE)
            meta       = insp_meta.get(insp_id, {})
            comp_dt    = meta.get("completion")
            date_txt   = comp_dt.strftime("%m/%d/%Y").lstrip("0").replace("/0", "/") if pd.notna(comp_dt) else ""
            inspector  = tex(meta.get("completed_by", ""))

            meta_parts = [rf"\textbf{{\textcolor{{navy}}{{\#{tex(str(insp_id))}}}}}", tex(location)]
            if date_txt:
                meta_parts.append(tex(date_txt))
            if inspector:
                meta_parts.append(inspector)
            meta_line  = r"\ \textbullet\ ".join(meta_parts)
            score_cell = rf"\textbf{{\textcolor[RGB]{{{score_rgb}}}{{{score_txt}}}}}"

            table_rows.append(
                rf"\rowcolor{{inspbg}}\multicolumn{{5}}{{>{{\raggedright\arraybackslash}}p{{{span_w}}}}}"
                rf"{{\small {meta_line}\hfill {score_cell}}} \\"
                r"\noalign{\hrule\penalty10000}"
            )

            # Photo strip if images exist
            imgs = insp_images.get(str(insp_id), [])
            if imgs:
                img_cells = []
                for i, img_bytes in enumerate(imgs[:6]):  # cap at 6 per inspection
                    img_file = tmp_dir / f"img_{insp_id}_{i}.jpg"
                    img_file.write_bytes(img_bytes)
                    img_cells.append(rf"\includegraphics[height=2cm,keepaspectratio]{{{fwdslash(img_file)}}}")
                photos_tex = r"\quad ".join(img_cells)
                table_rows.append(
                    rf"\rowcolor{{photobg}}\multicolumn{{5}}{{>{{\raggedright\arraybackslash}}p{{{span_w}}}}}{{{photos_tex}}} \\"
                    r"\noalign{\hrule\penalty10000}"
                )

        table_rows.append(
            rf"{tex(zone)} & {tex(space)} & {tex(element)} & {rating} & {tex(comment)} \\\hline"
        )

    table_body = "\n".join(table_rows)

    # --- Logo ---
    if logo_path and Path(logo_path).exists():
        logo_tex = rf"\includegraphics[height=1cm,keepaspectratio]{{{fwdslash(logo_path)}}}"
    else:
        logo_tex = rf"\textbf{{\textcolor{{navy}}{{{company_name}}}}}"

    # --- Full document ---
    return rf"""
\documentclass[10pt,letterpaper]{{article}}

% ---------- packages ----------
\usepackage{{fontspec}}
\usepackage{{geometry}}
\usepackage{{xcolor}}
\usepackage{{graphicx}}
\usepackage{{tikz}}
\usepackage{{calc}}
\usepackage{{xltabular}}
\usepackage{{booktabs}}
\usepackage{{array}}
\usepackage{{colortbl}}
\usepackage{{fancyhdr}}
\usepackage{{textcase}}
\usepackage{{microtype}}
\usepackage{{parskip}}
\usepackage{{multirow}}
\usepackage{{lastpage}}

% ---------- font ----------
\setmainfont{{Segoe UI}}[
  BoldFont      = Segoe UI Bold,
  ItalicFont    = Segoe UI Italic,
  Ligatures     = TeX
]

% ---------- page layout ----------
\geometry{{letterpaper, top=2.5cm, bottom=2cm, left=2cm, right=2cm,
           headheight=1.4cm, headsep=0.4cm}}

% ---------- brand colours ----------
\definecolor{{navy}}{{RGB}}{{37,64,143}}
\definecolor{{sky}}{{RGB}}{{0,176,240}}
\definecolor{{statbg}}{{RGB}}{{240,243,246}}
\definecolor{{slate}}{{RGB}}{{77,91,130}}
\definecolor{{inspbg}}{{RGB}}{{221,227,240}}
\definecolor{{photobg}}{{RGB}}{{247,248,252}}
\definecolor{{theadrow}}{{RGB}}{{37,64,143}}
\definecolor{{tablerule}}{{RGB}}{{180,186,200}}
\definecolor{{rowodd}}{{RGB}}{{255,255,255}}
\definecolor{{roweven}}{{RGB}}{{240,243,246}}
\definecolor{{bargreen}}{{HTML}}{{85CF5F}}
\definecolor{{barambr}}{{HTML}}{{F0B557}}
\definecolor{{barred}}{{HTML}}{{F05773}}
\newlength{{\barw}}

% ---------- running header (every page) ----------
\pagestyle{{fancy}}
\fancyhf{{}}
\renewcommand{{\headrulewidth}}{{0.4pt}}
\fancyhead[L]{{{logo_tex}}}
\fancyhead[C]{{\small\textcolor{{navy}}{{\textbf{{{venue}}}}}}}
\fancyhead[R]{{\small\textcolor{{slate}}{{{date_str}}}}}
\fancyfoot[R]{{\small\thepage\ of \pageref{{LastPage}}}}

% ---------- table helpers ----------
\newcolumntype{{L}}[1]{{>{{\raggedright\arraybackslash}}p{{#1}}}}
\newcolumntype{{R}}[1]{{>{{\raggedleft\arraybackslash}}p{{#1}}}}
\setlength{{\parindent}}{{0pt}}
\setlength{{\tabcolsep}}{{4pt}}
\renewcommand{{\arraystretch}}{{1.3}}
\setlength{{\arrayrulewidth}}{{0.4pt}}

\begin{{document}}

% ══════════════════════════════════════════
%  COVER / SUMMARY SECTION
% ══════════════════════════════════════════
\thispagestyle{{fancy}}

\vspace*{{0.4cm}}
% Score hero + title row
\noindent
\begin{{minipage}}[c]{{0.22\textwidth}}
  \centering
  \colorbox{{navy}}{{\parbox{{\dimexpr\linewidth-2\fboxsep}}{{
    \vspace{{6pt}}
    \centering
    \textbf{{\small\textcolor{{white}}{{\MakeUppercase{{Overall Score}}}}}}\\[4pt]
    \textbf{{\Huge\textcolor{{white}}{{{overall}\%}}}}
    \vspace{{6pt}}
  }}}}
\end{{minipage}}%
\hfill
\begin{{minipage}}[c]{{0.74\textwidth}}
  {{\large\textbf{{\textcolor{{navy}}{{Facility Inspection Report}}}}}}\\[2pt]
  {{\small\textcolor{{slate}}{{{venue}}}}}\\[1pt]
  {{\small\textcolor{{slate}}{{{date_str}}}}}
\end{{minipage}}

\vspace{{0.6cm}}

% Stat cards
\noindent
{cards}

{highlights}

\newpage

% Charts
{zone_chart}

\vspace{{0.5cm}}
{loc_chart}

\vspace{{0.5cm}}
{elem_chart}

\newpage

% ══════════════════════════════════════════
%  DEFICIENCY WORK ORDERS TABLE
% ══════════════════════════════════════════
\noindent{{\large\textbf{{\textcolor{{navy}}{{Deficiency Work Orders}}}}}}
\vspace{{0.3cm}}

% Fixed-width longtable (see COL_W in build_latex). \LTleft/\LTright pinned to
% 0pt so longtable doesn't centre the table.
\setlength\LTleft{{0pt}}\setlength\LTright{{0pt}}
\noindent\begin{{longtable}}{{{colspec}}}
{thead}
{table_body}
\end{{longtable}}

\end{{document}}
"""


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate_report(input_file: str, output_pdf: str = None, log=print,
                    comments: dict = None) -> str:
    """Generate a PDF report from an Excel inspection export.

    Args:
        input_file: Path to the .xlsx file.
        output_pdf:  Destination PDF path. Defaults to <input>_report.pdf.
        log:         Callable for progress messages (default: print).
        comments:    Optional dict with keys 'client', 'maintenance', 'general'.

    Returns:
        Absolute path to the generated PDF.
    """
    if output_pdf is None:
        output_pdf = str(Path(input_file).with_suffix("")) + "_report.pdf"

    xelatex = find_xelatex()

    log(f"[1] Reading {Path(input_file).name} ...")
    df, filters, raw_df = load_inspections(input_file)
    log(f"    {len(df)} complete rows across {df['Inspection #'].nunique()} inspections.")

    log("[2] Building summary ...")
    summary     = build_summary(df, filters)
    insp_images = extract_images(input_file, raw_df)
    if not insp_images:
        links = extract_image_links(input_file)
        n_links = sum(len(v) for v in links.values())
        if n_links:
            log(f"    No embedded images; downloading {n_links} linked images ...")
            insp_images = download_images(links, log=log)
    log(f"    Overall score : {summary['overall_score']}%")
    log(f"    Work orders   : {len(summary['work_order_rows'])}")
    log(f"    Images        : {sum(len(v) for v in insp_images.values())}")

    with tempfile.TemporaryDirectory() as tmp:
        tmp_dir = Path(tmp)

        log("[3] Rendering charts ...")
        zone_chart = render_bar_chart_tex(summary["zone_scores"],    "Performance Scores by Zone")
        loc_chart  = render_bar_chart_tex(summary["loc_scores"],     "Performance Scores by Location Type")
        elem_chart = render_bar_chart_tex(summary["element_scores"], "Performance Scores by Element")

        log("[4] Building LaTeX document ...")
        latex_src = build_latex(summary, zone_chart, loc_chart, elem_chart, insp_images, tmp_dir, comments=comments)

        tex_file = tmp_dir / "report.tex"
        tex_file.write_text(latex_src, encoding="utf-8")

        log("[5] Compiling PDF (pass 1) ...")
        subprocess.run(
            [xelatex, "-interaction=nonstopmode", "-output-directory", str(tmp_dir), str(tex_file)],
            capture_output=True, text=True
        )
        pdf_tmp = tmp_dir / "report.pdf"
        if not pdf_tmp.exists():
            log_file = tmp_dir / "report.log"
            if log_file.exists():
                lines = log_file.read_text(encoding="utf-8", errors="replace").splitlines()
                log("\n".join(lines[-30:]))
            raise RuntimeError("XeLaTeX failed — PDF not produced.")

        log("[5] Compiling PDF (pass 2) ...")
        subprocess.run(
            [xelatex, "-interaction=nonstopmode", "-output-directory", str(tmp_dir), str(tex_file)],
            capture_output=True, text=True
        )

        shutil.copy(pdf_tmp, output_pdf)

    log(f"[Done] PDF saved to: {output_pdf}")
    return str(Path(output_pdf).resolve())


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Generate facility inspection PDF report.")
    parser.add_argument("input",  help="Path to inspection .xlsx file")
    parser.add_argument("output", nargs="?", help="Output PDF path (default: <input>_report.pdf)")
    parser.add_argument("--comments", metavar="FILE",
                        help="JSON file with keys 'client', 'maintenance', 'general'")
    args = parser.parse_args()

    comments = {}
    if args.comments:
        with open(args.comments, encoding="utf-8") as f:
            comments = json.load(f)

    generate_report(args.input, args.output, comments=comments)


if __name__ == "__main__":
    main()
